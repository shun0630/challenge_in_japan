import sys
import openpyxl
import os
import shutil
from PIL import Image, ImageDraw, ImageFont
import numpy as np
import textwrap
import json


#エクセルからデータを読み込む
def input_data(data):
    file_name = data["exel_name"]
    wb = openpyxl.load_workbook(file_name, data_only=True)
    sheetnames_list = wb.sheetnames
    for sheetname in sheetnames_list:
        sheet = wb.get_sheet_by_name(str(sheetname))
        write_data = []
        folder_data = []

        for row in range(5,sheet.max_row+1):
            data_input = {}
            for item in data["row_data"]:
                data_input[item] = sheet[str(data["row_data"][item]) + str(row)].value
            print(data_input)
            print(data_input[data["folder_name"]], data_input[data["file_name"]])
            data_input["output_path"] = "./output/" + str(data_input[data["folder_name"]]) + "/" + str(data_input[data["file_name"]]) + ".png"
            write_data.append(data_input)
            folder_data.append(data_input[data["folder_name"]])
    folder_data_list = set(folder_data)
    print(folder_data_list)
    team_data_list = {"write_data":write_data, "folder_data":folder_data_list}
    return team_data_list
    

#団体ごとのフォルダを作成   
def mkdir_group(folder_data_list):
    shutil.rmtree('./output')
    os.mkdir("./output")
    for group in folder_data_list:
        os.mkdir('./output/' + str(group))




#フォントの設定
def font_setting(font_data, team_data, im):
    font_data_list = {}
    for data in font_data:
        font_data_list[data] =  font_position(team_data[data], font_data[data]["size"], font_data[data]["height"], font_data[data]["width"], im)
    return font_data_list



#書き込み場所、フォントサイズの設定
def font_position(write_comment, font_size, font_height_input, font_width_input, im):
    im_copy = im
    draw = ImageDraw.Draw(im_copy)
    font_path = "/usr/share/fonts/truetype/fonts-japanese-mincho.ttf"
    W_im, H_im = im.size
    font = ImageFont.truetype(font_path, int(font_size))
    w_text, h_text = draw.textsize(str(write_comment), font)
    print(w_text, h_text)
    
    #文字数に合わせてフォントサイズを調整する。
    while W_im - w_text < W_im/1.7:
        font_size = font_size - 5
        font = ImageFont.truetype(font_path, int(font_size))
        w_text, h_text = draw.textsize(str(write_comment), font)
    font_height = font_height_input
    font_width = W_im/2
    return {"write_comment": write_comment,"font": font, "font_height":font_height, "font_width": font_width}


#画像にテキストを書き込み
def image_write(data, im, font_data_list):
    im_copy = im
    draw = ImageDraw.Draw(im_copy)
    for write_item in font_data_list:
        draw.text((font_data_list[write_item]["font_width"], font_data_list[write_item]["font_height"]), str(data[write_item]), fill = "black", font = font_data_list[write_item]["font"], anchor='mm')
    im_copy.save(data["output_path"])    



def main():

    image_file = sys.argv[1]
    font_data_json = sys.argv[2]
    ecxel_data_json = sys.argv[3]
    
    #excelデータ(どこに何が書かれているか)の読み込み
    with open(ecxel_data_json) as f:
        ecxel_data = json.load(f)

    #フォントパラメータの読み込み（サイズ等）
    with open(font_data_json) as f:
        font_data = json.load(f)

    # データをエクセルから読み込み
    team_data_list = input_data(ecxel_data)

    #団体ごとのフォルダ作成
    mkdir_group(team_data_list["folder_data"])

    #書き込み用画像の用意
    im = Image.open(image_file)

    for team_data in team_data_list["write_data"]:
        #書き込み用データのセット
        im = Image.open(image_file)

        #フォントの設定
        font_data_list = font_setting(font_data, team_data, im)

        # 参加賞画像に書き込み
        image_write(team_data, im, font_data_list)
if __name__ == "__main__":
    main()